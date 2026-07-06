#!/usr/bin/env python3
"""
U-assays-style XLSX output module.

Generates an Excel file matching the format of U-assays_Scott-QC.xlsx:
27 columns per primer pair, one sheet "Primer pairs".

Columns:
1. assay_id, 2. seq_id, 3. template_used, 4. assay, 5. left_primer_display,
6. right_primer_display, 7. left_tm_C, 8. right_tm_C, 9. product_size_bp,
10. c_total_tail, 11. c_total, 12. left_c_total, 13. right_c_total,
14. left_c_tail, 15. right_c_tail, 16. sense_meth_mismatch_score,
17. sense_unmeth_mismatch_score, 18. anti_meth_mismatch_score,
19. anti_unmeth_mismatch_score, 20. bowtie_passes_filter,
21. bowtie_intended_genome, 22. left_structure_mfe_kcal_mol,
23. right_structure_mfe_kcal_mol, 24. primer_dimer_prediction,
25. primer_dimer_end_min_dg, 26. common_variant_score, 27. mapping_error_note
"""

import os
import pandas as pd
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from methyl_panel.phase3_primer3_design import PrimerPair


# Column order matching U-assays reference
COLUMNS = [
    "assay_id",
    "seq_id",
    "template_used",
    "assay",
    "left_primer_display",
    "right_primer_display",
    "left_tm_C",
    "right_tm_C",
    "product_size_bp",
    "c_total_tail",
    "c_total",
    "left_c_total",
    "right_c_total",
    "left_c_tail",
    "right_c_tail",
    "sense_meth_mismatch_score",
    "sense_unmeth_mismatch_score",
    "anti_meth_mismatch_score",
    "anti_unmeth_mismatch_score",
    "bowtie_passes_filter",
    "bowtie_intended_genome",
    "left_structure_mfe_kcal_mol",
    "right_structure_mfe_kcal_mol",
    "primer_dimer_prediction",
    "primer_dimer_end_min_dg",
    "common_variant_score",
    "mapping_error_note",
]


def primer_pair_to_row(p: PrimerPair) -> dict:
    """Convert a PrimerPair to a row dict matching U-assays format."""
    # Determine assay type from template
    if p.template_used in ("SM", "AM"):
        assay = "M"
    elif p.template_used in ("SU", "AU"):
        assay = "U"
    else:
        assay = p.template_used

    return {
        "assay_id": p.assay_id,
        "seq_id": p.seq_id,
        "template_used": p.template_used,
        "assay": assay,
        "left_primer_display": p.left_primer_display,
        "right_primer_display": p.right_primer_display,
        "left_tm_C": round(p.left_tm, 2),
        "right_tm_C": round(p.right_tm, 2),
        "product_size_bp": p.product_size,
        "c_total_tail": p.c_total_tail,
        "c_total": p.c_total,
        "left_c_total": p.left_c_total,
        "right_c_total": p.right_c_total,
        "left_c_tail": p.left_c_tail,
        "right_c_tail": p.right_c_tail,
        "sense_meth_mismatch_score": p.sense_meth_mismatch_score,
        "sense_unmeth_mismatch_score": p.sense_unmeth_mismatch_score,
        "anti_meth_mismatch_score": p.anti_meth_mismatch_score,
        "anti_unmeth_mismatch_score": p.anti_unmeth_mismatch_score,
        "bowtie_passes_filter": p.bowtie_passes_filter,
        "bowtie_intended_genome": p.bowtie_intended_genome,
        "left_structure_mfe_kcal_mol": p.left_structure_mfe,
        "right_structure_mfe_kcal_mol": p.right_structure_mfe,
        "primer_dimer_prediction": p.primer_dimer_prediction,
        "primer_dimer_end_min_dg": p.primer_dimer_end_min_dg,
        "common_variant_score": p.common_variant_score,
        "mapping_error_note": p.mapping_error_note,
    }


def write_xlsx(primer_pairs: List[PrimerPair], output_path: str,
               sheet_name: str = "Primer pairs") -> str:
    """
    Write primer pairs to an U-assays-style Excel file.

    Args:
        primer_pairs: List of PrimerPair objects
        output_path: Output file path
        sheet_name: Excel sheet name

    Returns:
        Path to the written file
    """
    # Convert to rows
    rows = [primer_pair_to_row(p) for p in primer_pairs]
    df = pd.DataFrame(rows, columns=COLUMNS)

    # Create workbook with styling
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(sheet_name)

    # Header styling
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', text_rotation=45)
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            # Color-code pass/fail
            if col_name == "bowtie_passes_filter" and val is not None:
                if val:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Auto-filter and freeze
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Column widths
    col_widths = {
        "assay_id": 15, "seq_id": 15, "template_used": 12, "assay": 8,
        "left_primer_display": 28, "right_primer_display": 28,
        "left_tm_C": 10, "right_tm_C": 10, "product_size_bp": 12,
        "c_total_tail": 10, "c_total": 8, "left_c_total": 10, "right_c_total": 10,
        "left_c_tail": 8, "right_c_tail": 8,
        "sense_meth_mismatch_score": 12, "sense_unmeth_mismatch_score": 12,
        "anti_meth_mismatch_score": 12, "anti_unmeth_mismatch_score": 12,
        "bowtie_passes_filter": 10, "bowtie_intended_genome": 18,
        "left_structure_mfe_kcal_mol": 12, "right_structure_mfe_kcal_mol": 12,
        "primer_dimer_prediction": 20, "primer_dimer_end_min_dg": 12,
        "common_variant_score": 10, "mapping_error_note": 30,
    }
    for col_idx, col_name in enumerate(COLUMNS, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

    # Write to workspace first (for S3-backed paths)
    if output_path.startswith("/mnt/"):
        tmp_path = "/workspace/_tmp_xlsx.xlsx"
        wb.save(tmp_path)
        import shutil
        shutil.copy(tmp_path, output_path)
        os.unlink(tmp_path)
    else:
        wb.save(output_path)

    return output_path


def main():
    """CLI entry point."""
    import argparse
    parser = argparse.ArgumentParser(description="Write U-assays-style XLSX")
    parser.add_argument("--output", required=True, help="Output XLSX path")
    args = parser.parse_args()

    # Create dummy data for testing
    from methyl_panel.phase3_primer3_design import PrimerPair
    p = PrimerPair(
        assay_id="MONO_0001_M_001", seq_id="MONO_0001", cell_type_id="MONO",
        template_used="SM", left_primer="ATTTATTAATTATTAATTAT",
        right_primer="GTTGTTGTTGTTGTTGTTGT",
        left_primer_display="yTTTATTyATTATTyATTAT",
        right_primer_display="GTTGTTGTTGTTGTTGTTGT",
        left_start=0, left_len=20, right_start=80, right_len=20,
        left_tm=59.5, right_tm=60.2, product_size=100,
        c_total_tail=2, c_total=5, left_c_total=2, right_c_total=1,
        left_c_tail=1, right_c_tail=0,
        sense_meth_mismatch_score=0, sense_unmeth_mismatch_score=3,
        anti_meth_mismatch_score=2, anti_unmeth_mismatch_score=5,
        left_gc_percent=30.0, right_gc_percent=35.0,
        left_self_any=2.0, right_self_any=1.5,
        left_self_end=0.5, right_self_end=0.0,
        pair_compl_any=3.0, pair_compl_end=1.0,
        left_end_stability=-2.0, right_end_stability=-1.5,
        penalty=0.85,
        bowtie_passes_filter=True, bowtie_intended_genome="converted_methylated",
        left_structure_mfe=-0.8, right_structure_mfe=-1.2,
        primer_dimer_prediction="low", primer_dimer_end_min_dg=-0.5,
        common_variant_score=0, mapping_error_note="Unique mapping",
    )

    write_xlsx([p], args.output)
    print(f"Written to {args.output}")


if __name__ == "__main__":
    main()
