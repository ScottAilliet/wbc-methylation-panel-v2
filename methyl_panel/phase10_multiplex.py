#!/usr/bin/env python3
"""
Phase 10: Multiplexing compatibility filter.

Selects one assay per cell type from the per-cell-type pipeline runs,
ensuring the selected assays are mutually compatible for multiplexed dPCR.

Three compatibility criteria:
1. Cross-dimer ΔG — all 4 cross-orientations between two assays' primers
   must have ΔG >= -1.0 kcal/mol (same DimerDetective method as Step 7).
2. Tm matching — the Tm spread across all 4 primers in the two assays
   must be <= 2.0 °C for uniform annealing.
3. Amplicon size separation — product sizes must differ by >= 10 bp
   for gel/capillary electrophoresis distinction.

Selection uses a greedy-by-quality algorithm with bounded backtracking:
- Assays within each cell type are ranked by a composite quality score.
- Cell types are processed in order of fewest available assays first.
- The best-ranked assay is tried; if incompatible with already-selected
  assays, the next-ranked is tried, up to a backtracking depth of 3.

Quality score components:
- DMR cleanliness score (40%)
- Primer3 penalty, normalized (25%)
- Self-dimer risk tier (15%)
- Bowtie specificity (10%)
- Tm closeness to optimal (10%)
"""

import os
import json
import primer3
import warnings
import numpy as np
from dataclasses import dataclass, field, asdict
from typing import List, Optional, Dict, Tuple

warnings.filterwarnings("ignore", category=RuntimeWarning)


# ─────────────────────────────────────────────────────────────────────────────
# Salt conditions — must match phase7_dimer.py for consistency
# ─────────────────────────────────────────────────────────────────────────────
MV_CONC = 50.0      # mM monovalent
DV_CONC = 1.5       # mM divalent
DNTP_CONC = 0.6     # mM dNTP
DNA_CONC = 50.0     # nM primer
TEMP_C = 37.0       # °C

DEFAULT_CROSS_DIMER_CUTOFF = -1.0   # kcal/mol
DEFAULT_TM_TOLERANCE = 2.0          # °C
DEFAULT_MIN_AMPLICON_DIFF = 10      # bp
DEFAULT_BACKTRACK_DEPTH = 3


# ─────────────────────────────────────────────────────────────────────────────
# Data structures
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class CrossDimerResult:
    """Cross-dimer score between two primers from different assays."""
    dg: float               # ΔG in kcal/mol (most stable orientation)
    orientation: str        # Which orientation was worst
    passes: bool            # True if dg >= cutoff


@dataclass
class AssayCompatibility:
    """Pairwise compatibility between two assays from different cell types."""
    assay_a_id: str
    assay_b_id: str
    cell_type_a: str
    cell_type_b: str
    cross_dimer_dg: float       # Min ΔG across all 4 cross-orientations
    cross_dimer_worst_orientation: str
    cross_dimer_passes: bool
    tm_spread: float            # max(Tm) - min(Tm) across all 4 primers
    tm_passes: bool
    amplicon_diff: int          # |product_size_A - product_size_B|
    amplicon_passes: bool
    compatible: bool            # All three criteria pass


@dataclass
class MultiplexPanel:
    """Selected multiplex panel with full provenance."""
    selected_assays: List[dict] = field(default_factory=list)
    compatibility_matrix: List[dict] = field(default_factory=list)
    selection_log: List[str] = field(default_factory=list)
    failed_cell_types: List[str] = field(default_factory=list)
    n_cell_types_requested: int = 0
    n_cell_types_selected: int = 0
    criteria: dict = field(default_factory=dict)


# ─────────────────────────────────────────────────────────────────────────────
# Cross-dimer scoring
# ─────────────────────────────────────────────────────────────────────────────

def _calc_end_stability_dg(seq1: str, seq2: str) -> float:
    """
    Calculate 3' end stability (ΔG) between two sequences.
    Same method as phase7_dimer.calc_end_stability_dg.
    Returns ΔG in kcal/mol (negative = stable = dimer risk).
    """
    try:
        result = primer3.calc_end_stability(
            seq1, seq2,
            mv_conc=MV_CONC,
            dv_conc=DV_CONC,
            dntp_conc=DNTP_CONC,
            dna_conc=DNA_CONC,
            temp_c=TEMP_C,
        )
        return result.dg / 1000.0  # cal/mol → kcal/mol
    except Exception:
        return 0.0


def calc_cross_dimer(assay_a: dict, assay_b: dict,
                     cutoff: float = DEFAULT_CROSS_DIMER_CUTOFF) -> CrossDimerResult:
    """
    Calculate cross-dimer risk between two assays.

    Evaluates all 8 orientations: the 3' end of each of the 4 primers
    (A_left, A_right, B_left, B_right) against each of the 2 primers
    in the *other* assay. The minimum ΔG determines the risk.

    primer3.calc_end_stability is directional (evaluates 3' end of seq1
    against seq2), so A_left→B_left ≠ B_left→A_left. We compute all 8:
      A_left  3' end → B_left,  A_left  3' end → B_right
      A_right 3' end → B_left,  A_right 3' end → B_right
      B_left  3' end → A_left,  B_left  3' end → A_right
      B_right 3' end → A_left,  B_right 3' end → A_right

    Args:
        assay_a: First assay dict (must have left_primer, right_primer)
        assay_b: Second assay dict (must have left_primer, right_primer)
        cutoff: ΔG cutoff (kcal/mol). Pass if min ΔG >= cutoff.

    Returns:
        CrossDimerResult with worst ΔG and orientation.
    """
    a_l = assay_a["left_primer"]
    a_r = assay_a["right_primer"]
    b_l = assay_b["left_primer"]
    b_r = assay_b["right_primer"]

    orientations = {
        "A_left→B_left":   _calc_end_stability_dg(a_l, b_l),
        "A_left→B_right":  _calc_end_stability_dg(a_l, b_r),
        "A_right→B_left":  _calc_end_stability_dg(a_r, b_l),
        "A_right→B_right": _calc_end_stability_dg(a_r, b_r),
        "B_left→A_left":   _calc_end_stability_dg(b_l, a_l),
        "B_left→A_right":  _calc_end_stability_dg(b_l, a_r),
        "B_right→A_left":  _calc_end_stability_dg(b_r, a_l),
        "B_right→A_right": _calc_end_stability_dg(b_r, a_r),
    }

    worst_orient = min(orientations, key=orientations.get)
    worst_dg = orientations[worst_orient]

    return CrossDimerResult(
        dg=round(worst_dg, 2),
        orientation=worst_orient,
        passes=worst_dg >= cutoff,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Pairwise compatibility
# ─────────────────────────────────────────────────────────────────────────────

def check_compatibility(assay_a: dict, assay_b: dict,
                        cross_dimer_cutoff: float = DEFAULT_CROSS_DIMER_CUTOFF,
                        tm_tolerance: float = DEFAULT_TM_TOLERANCE,
                        min_amplicon_diff: int = DEFAULT_MIN_AMPLICON_DIFF
                        ) -> AssayCompatibility:
    """
    Check all three compatibility criteria between two assays.

    Args:
        assay_a, assay_b: Assay dicts from primers.json
        cross_dimer_cutoff: Max allowed ΔG stability (kcal/mol)
        tm_tolerance: Max Tm spread across all 4 primers (°C)
        min_amplicon_diff: Min product size difference (bp)

    Returns:
        AssayCompatibility with all criteria evaluated.
    """
    # 1. Cross-dimer
    xd = calc_cross_dimer(assay_a, assay_b, cutoff=cross_dimer_cutoff)

    # 2. Tm matching — spread across all 4 primers
    tms = [
        assay_a["left_tm"], assay_a["right_tm"],
        assay_b["left_tm"], assay_b["right_tm"],
    ]
    tm_spread = round(max(tms) - min(tms), 2)
    tm_passes = tm_spread <= tm_tolerance

    # 3. Amplicon size separation
    amplicon_diff = abs(assay_a["product_size"] - assay_b["product_size"])
    amplicon_passes = amplicon_diff >= min_amplicon_diff

    return AssayCompatibility(
        assay_a_id=assay_a["assay_id"],
        assay_b_id=assay_b["assay_id"],
        cell_type_a=assay_a.get("cell_type_id", "?"),
        cell_type_b=assay_b.get("cell_type_id", "?"),
        cross_dimer_dg=xd.dg,
        cross_dimer_worst_orientation=xd.orientation,
        cross_dimer_passes=xd.passes,
        tm_spread=tm_spread,
        tm_passes=tm_passes,
        amplicon_diff=amplicon_diff,
        amplicon_passes=amplicon_passes,
        compatible=xd.passes and tm_passes and amplicon_passes,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Quality ranking
# ─────────────────────────────────────────────────────────────────────────────

def compute_quality_score(assay: dict, cleanliness: Optional[float],
                          opt_tm: float = 60.0,
                          max_penalty: float = 1.0) -> float:
    """
    Compute composite quality score for an assay.

    Components:
        cleanliness_score (40%) — DMR quality, 0-1
        primer3 penalty (25%) — normalized, lower = better
        self-dimer tier (15%) — low=1.0, medium=0.5, high=0.0
        bowtie specificity (10%) — pass=1.0, fail/None=0.0
        Tm closeness (10%) — how close both Tms are to opt_tm

    Args:
        assay: Assay dict from primers.json
        cleanliness: Cleanliness score from dmr_blocks.json (or None)
        opt_tm: Optimal Tm for scoring
        max_penalty: Max penalty across cell type for normalization

    Returns:
        Quality score (0-1, higher = better)
    """
    # Cleanliness (0-1)
    cl = cleanliness if cleanliness is not None else 0.5

    # Penalty (lower = better → invert)
    penalty = assay.get("penalty", 1.0)
    norm_penalty = min(penalty / max(max_penalty, 0.001), 1.0)
    penalty_score = 1.0 - norm_penalty

    # Self-dimer tier
    dimer_pred = assay.get("primer_dimer_prediction", "low")
    if dimer_pred == "low":
        dimer_score = 1.0
    elif dimer_pred == "medium":
        dimer_score = 0.5
    else:
        dimer_score = 0.0

    # Bowtie specificity
    bowtie = assay.get("bowtie_passes_filter")
    bowtie_score = 1.0 if bowtie is True else 0.0

    # Tm closeness (average of both primers)
    left_tm_closeness = max(0.0, 1.0 - abs(assay["left_tm"] - opt_tm) / 5.0)
    right_tm_closeness = max(0.0, 1.0 - abs(assay["right_tm"] - opt_tm) / 5.0)
    tm_score = (left_tm_closeness + right_tm_closeness) / 2.0

    score = (
        cl * 0.40
        + penalty_score * 0.25
        + dimer_score * 0.15
        + bowtie_score * 0.10
        + tm_score * 0.10
    )

    return round(score, 4)


# ─────────────────────────────────────────────────────────────────────────────
# Data loading
# ─────────────────────────────────────────────────────────────────────────────

def load_cell_type_data(output_dir: str) -> Tuple[List[dict], Dict[str, float]]:
    """
    Load primers.json and dmr_blocks.json from a cell-type output directory.

    Args:
        output_dir: Path to the per-cell-type output directory

    Returns:
        (primers_list, cleanliness_map) where cleanliness_map maps
        seq_id → cleanliness_score
    """
    primers_path = os.path.join(output_dir, "primers.json")
    if not os.path.exists(primers_path):
        raise FileNotFoundError(f"No primers.json in {output_dir}")

    with open(primers_path) as f:
        primers = json.load(f)

    # Load cleanliness scores from dmr_blocks.json
    cleanliness_map = {}
    blocks_path = os.path.join(output_dir, "dmr_blocks.json")
    if os.path.exists(blocks_path):
        with open(blocks_path) as f:
            blocks = json.load(f)
        for b in blocks:
            cleanliness_map[b["seq_id"]] = b.get("cleanliness_score", 0.5)

    return primers, cleanliness_map


def prepare_candidate_assays(primers: List[dict],
                             cleanliness_map: Dict[str, float],
                             opt_tm: float = 60.0) -> List[dict]:
    """
    Rank assays by quality score and return sorted list.

    Each assay dict is augmented with:
        _quality_score: float
        _cleanliness: float or None

    Args:
        primers: List of assay dicts from primers.json
        cleanliness_map: seq_id → cleanliness_score
        opt_tm: Optimal Tm for scoring

    Returns:
        List of assay dicts sorted by quality score (descending)
    """
    if not primers:
        return []

    # Compute max penalty for normalization
    max_penalty = max(p.get("penalty", 1.0) for p in primers)

    for p in primers:
        cl = cleanliness_map.get(p.get("seq_id", ""))
        p["_quality_score"] = compute_quality_score(p, cl, opt_tm, max_penalty)
        p["_cleanliness"] = cl

    primers.sort(key=lambda p: p["_quality_score"], reverse=True)
    return primers


# ─────────────────────────────────────────────────────────────────────────────
# Greedy selection with backtracking
# ─────────────────────────────────────────────────────────────────────────────

def select_multiplex_panel(
    cell_type_dirs: Dict[str, str],
    opt_tm: float = 60.0,
    cross_dimer_cutoff: float = DEFAULT_CROSS_DIMER_CUTOFF,
    tm_tolerance: float = DEFAULT_TM_TOLERANCE,
    min_amplicon_diff: int = DEFAULT_MIN_AMPLICON_DIFF,
    backtrack_depth: int = DEFAULT_BACKTRACK_DEPTH,
) -> MultiplexPanel:
    """
    Select one assay per cell type for a multiplex panel.

    Uses greedy-by-quality selection with bounded backtracking:
    1. Load and rank assays for each cell type
    2. Process cell types in order of fewest candidates first
    3. Try best-ranked assay; check compatibility with selected
    4. If incompatible, try next; backtrack if exhausted

    Args:
        cell_type_dirs: {cell_type_id: output_dir_path}
        opt_tm: Optimal Tm for quality scoring
        cross_dimer_cutoff: Max ΔG for cross-dimer (kcal/mol)
        tm_tolerance: Max Tm spread (°C)
        min_amplicon_diff: Min product size difference (bp)
        backtrack_depth: Max backtracking depth

    Returns:
        MultiplexPanel with selected assays, compatibility matrix,
        selection log, and any failed cell types.
    """
    log = []
    panel = MultiplexPanel(
        n_cell_types_requested=len(cell_type_dirs),
        criteria={
            "cross_dimer_cutoff": cross_dimer_cutoff,
            "tm_tolerance": tm_tolerance,
            "min_amplicon_diff": min_amplicon_diff,
            "opt_tm": opt_tm,
        },
    )

    # ── Load and rank assays for each cell type ──
    candidates: Dict[str, List[dict]] = {}
    for ct_id, out_dir in cell_type_dirs.items():
        try:
            primers, cl_map = load_cell_type_data(out_dir)
            ranked = prepare_candidate_assays(primers, cl_map, opt_tm)
            candidates[ct_id] = ranked
            log.append(f"Loaded {len(ranked)} assays for {ct_id} from {out_dir}")
        except FileNotFoundError as e:
            log.append(f"WARNING: {e}")
            candidates[ct_id] = []

    # ── Order cell types by number of candidates (ascending) ──
    ct_order = sorted(candidates.keys(), key=lambda ct: len(candidates[ct]))

    # ── Greedy selection with backtracking ──
    # State: list of (cell_type, assay_index) tuples
    selected: List[Tuple[str, dict]] = []
    # Track which index we're trying for each cell type
    indices: Dict[str, int] = {}

    def _check_against_selected(assay: dict) -> bool:
        """Check if assay is compatible with all currently selected assays."""
        for _, sel_assay in selected:
            compat = check_compatibility(
                assay, sel_assay,
                cross_dimer_cutoff=cross_dimer_cutoff,
                tm_tolerance=tm_tolerance,
                min_amplicon_diff=min_amplicon_diff,
            )
            if not compat.compatible:
                log.append(
                    f"    ✗ {assay['assay_id']} incompatible with "
                    f"{sel_assay['assay_id']} "
                    f"(dimer={compat.cross_dimer_dg}, "
                    f"tm_spread={compat.tm_spread}, "
                    f"amp_diff={compat.amplicon_diff})"
                )
                return False
        return True

    ct_idx = 0
    while ct_idx < len(ct_order):
        ct = ct_order[ct_idx]
        ct_candidates = candidates[ct]

        if not ct_candidates:
            log.append(f"  {ct}: no candidates available — marking as failed")
            panel.failed_cell_types.append(ct)
            ct_idx += 1
            continue

        # Try to find a compatible assay for this cell type
        start_idx = indices.get(ct, 0)
        found = False

        for i in range(start_idx, len(ct_candidates)):
            assay = ct_candidates[i]
            indices[ct] = i + 1  # Next time, start from i+1

            log.append(
                f"  {ct}: trying #{i+1}/{len(ct_candidates)} "
                f"{assay['assay_id']} (q={assay['_quality_score']:.3f})"
            )

            if _check_against_selected(assay):
                selected.append((ct, assay))
                log.append(f"    ✓ SELECTED {assay['assay_id']}")
                found = True
                break

        if found:
            ct_idx += 1
            continue

        # No compatible assay found — backtrack
        if selected and backtrack_depth > 0:
            # Remove last selected, try its next option
            prev_ct, prev_assay = selected.pop()
            log.append(
                f"  BACKTRACK: removing {prev_ct}/{prev_assay['assay_id']}, "
                f"retrying from index {indices.get(prev_ct, 0)}"
            )
            ct_idx -= 1
            # Continue loop — will retry prev_ct from indices[prev_ct]
        else:
            # Can't backtrack further — mark as failed
            log.append(f"  {ct}: no compatible assay found after backtracking — FAILED")
            panel.failed_cell_types.append(ct)
            # Reset indices for this CT and move on
            indices[ct] = 0
            ct_idx += 1

    # ── Build compatibility matrix for selected assays ──
    compat_matrix = []
    for i in range(len(selected)):
        for j in range(i + 1, len(selected)):
            ct_a, assay_a = selected[i]
            ct_b, assay_b = selected[j]
            compat = check_compatibility(
                assay_a, assay_b,
                cross_dimer_cutoff=cross_dimer_cutoff,
                tm_tolerance=tm_tolerance,
                min_amplicon_diff=min_amplicon_diff,
            )
            compat_matrix.append(asdict(compat))

    # ── Build output ──
    panel.selected_assays = [
        {
            "cell_type_id": ct,
            "assay_id": a["assay_id"],
            "seq_id": a["seq_id"],
            "template_used": a["template_used"],
            "left_primer": a["left_primer"],
            "right_primer": a["right_primer"],
            "left_primer_display": a["left_primer_display"],
            "right_primer_display": a["right_primer_display"],
            "left_tm": a["left_tm"],
            "right_tm": a["right_tm"],
            "product_size": a["product_size"],
            "c_total_tail": a["c_total_tail"],
            "c_total": a["c_total"],
            "left_c_total": a["left_c_total"],
            "right_c_total": a["right_c_total"],
            "left_c_tail": a["left_c_tail"],
            "right_c_tail": a["right_c_tail"],
            "sense_meth_mismatch_score": a["sense_meth_mismatch_score"],
            "sense_unmeth_mismatch_score": a["sense_unmeth_mismatch_score"],
            "anti_meth_mismatch_score": a["anti_meth_mismatch_score"],
            "anti_unmeth_mismatch_score": a["anti_unmeth_mismatch_score"],
            "left_gc_percent": a["left_gc_percent"],
            "right_gc_percent": a["right_gc_percent"],
            "penalty": a["penalty"],
            "bowtie_passes_filter": a.get("bowtie_passes_filter"),
            "bowtie_intended_genome": a.get("bowtie_intended_genome"),
            "left_structure_mfe": a.get("left_structure_mfe"),
            "right_structure_mfe": a.get("right_structure_mfe"),
            "primer_dimer_prediction": a.get("primer_dimer_prediction"),
            "primer_dimer_end_min_dg": a.get("primer_dimer_end_min_dg"),
            "common_variant_score": a.get("common_variant_score"),
            "mapping_error_note": a.get("mapping_error_note"),
            "quality_score": a["_quality_score"],
            "cleanliness_score": a.get("_cleanliness"),
        }
        for ct, a in selected
    ]
    panel.compatibility_matrix = compat_matrix
    panel.selection_log = log
    panel.n_cell_types_selected = len(selected)

    return panel


# ─────────────────────────────────────────────────────────────────────────────
# Output: JSON
# ─────────────────────────────────────────────────────────────────────────────

def save_panel_json(panel: MultiplexPanel, output_path: str) -> str:
    """Save multiplex panel to JSON."""
    data = {
        "n_cell_types_requested": panel.n_cell_types_requested,
        "n_cell_types_selected": panel.n_cell_types_selected,
        "failed_cell_types": panel.failed_cell_types,
        "criteria": panel.criteria,
        "selected_assays": panel.selected_assays,
        "compatibility_matrix": panel.compatibility_matrix,
        "selection_log": panel.selection_log,
    }
    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# Output: XLSX
# ─────────────────────────────────────────────────────────────────────────────

def save_panel_xlsx(panel: MultiplexPanel, output_path: str) -> str:
    """
    Save multiplex panel to XLSX with two sheets:
    1. "Multiplex panel" — one row per selected assay (same columns as
       primer_assays.xlsx plus cell_type_id and quality_score)
    2. "Compatibility matrix" — pairwise compatibility between all
       selected assays
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Multiplex panel ──
    ws = wb.create_sheet("Multiplex panel")

    panel_columns = [
        "cell_type_id", "assay_id", "seq_id", "template_used",
        "left_primer_display", "right_primer_display",
        "left_tm", "right_tm", "product_size",
        "c_total_tail", "c_total", "left_c_total", "right_c_total",
        "left_c_tail", "right_c_tail",
        "sense_meth_mismatch_score", "sense_unmeth_mismatch_score",
        "anti_meth_mismatch_score", "anti_unmeth_mismatch_score",
        "bowtie_passes_filter", "bowtie_intended_genome",
        "left_structure_mfe", "right_structure_mfe",
        "primer_dimer_prediction", "primer_dimer_end_min_dg",
        "common_variant_score", "mapping_error_note",
        "quality_score", "cleanliness_score",
    ]

    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, col_name in enumerate(panel_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', text_rotation=45)
        cell.border = thin_border

    for row_idx, assay in enumerate(panel.selected_assays, 2):
        for col_idx, col_name in enumerate(panel_columns, 1):
            val = assay.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    col_widths = {
        "cell_type_id": 10, "assay_id": 18, "seq_id": 15,
        "template_used": 12, "left_primer_display": 28,
        "right_primer_display": 28, "left_tm": 10, "right_tm": 10,
        "product_size": 12, "c_total_tail": 10, "c_total": 8,
        "left_c_total": 10, "right_c_total": 10,
        "left_c_tail": 8, "right_c_tail": 8,
        "quality_score": 12, "cleanliness_score": 12,
    }
    from openpyxl.utils import get_column_letter
    for col_idx, col_name in enumerate(panel_columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

    # ── Sheet 2: Compatibility matrix ──
    ws2 = wb.create_sheet("Compatibility matrix")

    compat_columns = [
        "cell_type_a", "assay_a_id", "cell_type_b", "assay_b_id",
        "cross_dimer_dg", "cross_dimer_worst_orientation", "cross_dimer_passes",
        "tm_spread", "tm_passes",
        "amplicon_diff", "amplicon_passes",
        "compatible",
    ]

    for col_idx, col_name in enumerate(compat_columns, 1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', text_rotation=45)
        cell.border = thin_border

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx, compat in enumerate(panel.compatibility_matrix, 2):
        for col_idx, col_name in enumerate(compat_columns, 1):
            val = compat.get(col_name)
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if col_name == "compatible":
                cell.fill = green_fill if val else red_fill

    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = "A2"

    compat_widths = {
        "cell_type_a": 10, "assay_a_id": 18,
        "cell_type_b": 10, "assay_b_id": 18,
        "cross_dimer_dg": 14, "cross_dimer_worst_orientation": 22,
        "cross_dimer_passes": 10, "tm_spread": 10, "tm_passes": 10,
        "amplicon_diff": 12, "amplicon_passes": 12, "compatible": 12,
    }
    for col_idx, col_name in enumerate(compat_columns, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = compat_widths.get(col_name, 15)

    # ── Sheet 3: Selection log ──
    ws3 = wb.create_sheet("Selection log")
    ws3.cell(row=1, column=1, value="Step").font = header_font
    ws3.cell(row=1, column=1).fill = header_fill
    ws3.cell(row=1, column=2, value="Log entry").font = header_font
    ws3.cell(row=1, column=2).fill = header_fill
    for i, entry in enumerate(panel.selection_log, 2):
        ws3.cell(row=i, column=1, value=i - 1).alignment = Alignment(horizontal='center')
        ws3.cell(row=i, column=2, value=entry)
    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 100

    # ── Save (handle S3-backed paths) ──
    if output_path.startswith("/mnt/"):
        tmp_path = "/workspace/_tmp_multiplex.xlsx"
        wb.save(tmp_path)
        import shutil
        shutil.copy(tmp_path, output_path)
        os.unlink(tmp_path)
    else:
        wb.save(output_path)

    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# Output: PDF
# ─────────────────────────────────────────────────────────────────────────────

def save_panel_pdf(panel: MultiplexPanel, output_path: str) -> str:
    """
    Save multiplex panel to PDF:
    - Page 1: Summary (panel composition + compatibility matrix table)
    - Subsequent pages: One page per selected assay (same format as
      primer_assays.pdf, using output_pdf.generate_primer_page)
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak,
    )
    from methyl_panel.phase3_primer3_design import PrimerPair

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('MuxTitle', parent=styles['Title'],
                                 fontSize=16, spaceAfter=10)
    header_style = ParagraphStyle('MuxHeader', parent=styles['Heading2'],
                                  fontSize=12, spaceAfter=6,
                                  textColor=colors.HexColor('#4472C4'))
    normal_style = styles['Normal']

    elements = []

    # ── Summary page ──
    elements.append(Paragraph("Multiplex Panel Summary", title_style))
    elements.append(Spacer(1, 6))

    # Panel composition
    ct_list = [a["cell_type_id"] for a in panel.selected_assays]
    failed = panel.failed_cell_types
    elements.append(Paragraph(
        f"Cell types selected: {panel.n_cell_types_selected}/{panel.n_cell_types_requested}",
        normal_style
    ))
    elements.append(Paragraph(f"Selected: {', '.join(ct_list)}", normal_style))
    if failed:
        elements.append(Paragraph(
            f"<font color='red'>Failed (no compatible assay): {', '.join(failed)}</font>",
            normal_style
        ))
    elements.append(Spacer(1, 8))

    # Criteria
    elements.append(Paragraph("Selection Criteria", header_style))
    criteria_text = (
        f"Cross-dimer cutoff: {panel.criteria.get('cross_dimer_cutoff', -1.0)} kcal/mol<br/>"
        f"Tm tolerance: {panel.criteria.get('tm_tolerance', 2.0)} °C<br/>"
        f"Min amplicon difference: {panel.criteria.get('min_amplicon_diff', 10)} bp<br/>"
        f"Optimal Tm: {panel.criteria.get('opt_tm', 60.0)} °C"
    )
    elements.append(Paragraph(criteria_text, normal_style))
    elements.append(Spacer(1, 8))

    # Selected assays table
    elements.append(Paragraph("Selected Assays", header_style))
    assay_data = [["Cell type", "Assay ID", "Left Tm", "Right Tm", "Product", "Quality"]]
    for a in panel.selected_assays:
        assay_data.append([
            a["cell_type_id"], a["assay_id"],
            f"{a['left_tm']:.1f}", f"{a['right_tm']:.1f}",
            f"{a['product_size']} bp",
            f"{a['quality_score']:.3f}",
        ])
    assay_table = Table(assay_data, colWidths=[50, 90, 50, 50, 60, 50])
    assay_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(assay_table)
    elements.append(Spacer(1, 12))

    # Compatibility matrix
    elements.append(Paragraph("Pairwise Compatibility Matrix", header_style))
    compat_data = [["CT A", "Assay A", "CT B", "Assay B",
                    "Dimer ΔG", "Tm spread", "Amp diff", "Compatible"]]
    for c in panel.compatibility_matrix:
        compat_data.append([
            c["cell_type_a"], c["assay_a_id"],
            c["cell_type_b"], c["assay_b_id"],
            f"{c['cross_dimer_dg']:.2f}",
            f"{c['tm_spread']:.2f}",
            f"{c['amplicon_diff']}",
            "✓" if c["compatible"] else "✗",
        ])
    compat_table = Table(compat_data, colWidths=[35, 75, 35, 75, 45, 45, 40, 50])
    compat_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]
    # Color-code compatible column
    for i, c in enumerate(panel.compatibility_matrix, 1):
        if c["compatible"]:
            compat_style.append(('BACKGROUND', (7, i), (7, i), colors.HexColor('#C6EFCE')))
        else:
            compat_style.append(('BACKGROUND', (7, i), (7, i), colors.HexColor('#FFC7CE')))
    compat_table.setStyle(TableStyle(compat_style))
    elements.append(compat_table)

    # ── One page per selected assay ──
    from methyl_panel.output_pdf import generate_primer_page

    for a in panel.selected_assays:
        elements.append(PageBreak())

        # Reconstruct PrimerPair for the page generator
        p = PrimerPair(
            assay_id=a["assay_id"], seq_id=a["seq_id"],
            cell_type_id=a["cell_type_id"], template_used=a["template_used"],
            left_primer=a["left_primer"], right_primer=a["right_primer"],
            left_primer_display=a["left_primer_display"],
            right_primer_display=a["right_primer_display"],
            left_start=0, left_len=len(a["left_primer"]),
            right_start=0, right_len=len(a["right_primer"]),
            left_tm=a["left_tm"], right_tm=a["right_tm"],
            product_size=a["product_size"],
            c_total_tail=a["c_total_tail"], c_total=a["c_total"],
            left_c_total=a["left_c_total"], right_c_total=a["right_c_total"],
            left_c_tail=a["left_c_tail"], right_c_tail=a["right_c_tail"],
            sense_meth_mismatch_score=a["sense_meth_mismatch_score"],
            sense_unmeth_mismatch_score=a["sense_unmeth_mismatch_score"],
            anti_meth_mismatch_score=a["anti_meth_mismatch_score"],
            anti_unmeth_mismatch_score=a["anti_unmeth_mismatch_score"],
            left_gc_percent=a["left_gc_percent"], right_gc_percent=a["right_gc_percent"],
            left_self_any=0, right_self_any=0, left_self_end=0, right_self_end=0,
            pair_compl_any=0, pair_compl_end=0,
            left_end_stability=0, right_end_stability=0,
            penalty=a["penalty"],
            bowtie_passes_filter=a.get("bowtie_passes_filter"),
            bowtie_intended_genome=a.get("bowtie_intended_genome"),
            left_structure_mfe=a.get("left_structure_mfe"),
            right_structure_mfe=a.get("right_structure_mfe"),
            primer_dimer_prediction=a.get("primer_dimer_prediction"),
            primer_dimer_end_min_dg=a.get("primer_dimer_end_min_dg"),
            common_variant_score=a.get("common_variant_score"),
            mapping_error_note=a.get("mapping_error_note"),
        )

        page_elements = generate_primer_page(p, strands=None)
        elements.extend(page_elements)

    # Build PDF
    if output_path.startswith("/mnt/"):
        tmp_path = "/workspace/_tmp_multiplex.pdf"
        doc = SimpleDocTemplate(tmp_path, pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)
        doc.build(elements)
        import shutil
        shutil.copy(tmp_path, output_path)
        os.unlink(tmp_path)
    else:
        doc = SimpleDocTemplate(output_path, pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)
        doc.build(elements)

    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    """CLI entry point for standalone execution."""
    import argparse
    parser = argparse.ArgumentParser(
        description="Multiplexing compatibility filter — select one assay per cell type"
    )
    parser.add_argument("--multiplex-dirs", required=True,
                        help="Comma-separated list of per-cell-type output dirs "
                             "(e.g. results/MONO,results/BCELL,...)")
    parser.add_argument("--output-dir", default="results/multiplex/",
                        help="Output directory for multiplex panel")
    parser.add_argument("--opt-tm", type=float, default=60.0,
                        help="Optimal Tm for quality scoring (°C)")
    parser.add_argument("--cross-dimer-cutoff", type=float,
                        default=DEFAULT_CROSS_DIMER_CUTOFF,
                        help=f"Cross-dimer ΔG cutoff (kcal/mol, default {DEFAULT_CROSS_DIMER_CUTOFF})")
    parser.add_argument("--tm-tolerance", type=float,
                        default=DEFAULT_TM_TOLERANCE,
                        help=f"Max Tm spread across compatible assays (°C, default {DEFAULT_TM_TOLERANCE})")
    parser.add_argument("--min-amplicon-diff", type=int,
                        default=DEFAULT_MIN_AMPLICON_DIFF,
                        help=f"Min product size difference (bp, default {DEFAULT_MIN_AMPLICON_DIFF})")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    # Parse multiplex dirs — auto-detect cell type from directory name
    dirs = [d.strip() for d in args.multiplex_dirs.split(",")]
    cell_type_dirs = {}
    for d in dirs:
        ct = os.path.basename(os.path.normpath(d))
        cell_type_dirs[ct] = d

    print(f"\n=== Step 10: Multiplexing Compatibility Filter ===")
    print(f"  Cell types: {list(cell_type_dirs.keys())}")
    print(f"  Criteria: dimer cutoff={args.cross_dimer_cutoff}, "
          f"tm tol={args.tm_tolerance}, min amp diff={args.min_amplicon_diff}")

    panel = select_multiplex_panel(
        cell_type_dirs,
        opt_tm=args.opt_tm,
        cross_dimer_cutoff=args.cross_dimer_cutoff,
        tm_tolerance=args.tm_tolerance,
        min_amplicon_diff=args.min_amplicon_diff,
    )

    # Save outputs
    json_path = save_panel_json(panel, os.path.join(args.output_dir, "multiplex_panel.json"))
    print(f"\n  Panel JSON: {json_path}")

    xlsx_path = save_panel_xlsx(panel, os.path.join(args.output_dir, "multiplex_panel.xlsx"))
    print(f"  Panel XLSX: {xlsx_path}")

    pdf_path = save_panel_pdf(panel, os.path.join(args.output_dir, "multiplex_panel.pdf"))
    print(f"  Panel PDF: {pdf_path}")

    # Summary
    print(f"\n  Selected: {panel.n_cell_types_selected}/{panel.n_cell_types_requested} cell types")
    if panel.failed_cell_types:
        print(f"  Failed: {', '.join(panel.failed_cell_types)}")
    for a in panel.selected_assays:
        print(f"    {a['cell_type_id']}: {a['assay_id']} "
              f"(q={a['quality_score']:.3f}, Tm={a['left_tm']:.1f}/{a['right_tm']:.1f}, "
              f"product={a['product_size']}bp)")


if __name__ == "__main__":
    main()
