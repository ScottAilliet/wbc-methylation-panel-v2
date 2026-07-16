#!/usr/bin/env python3
"""
Generate a master Excel file combining all primer assays across all cell types.

Scans results/*/ directories for primers.json and dmr_blocks.json,
joins them on seq_id, and writes a single XLSX with multiple sheets:

  - All_Primers:    Every primer pair (~4,335 rows), enriched with DMR-level info
  - Best_Per_DMR:   Top primer per DMR block (~760 rows), ranked by bowtie pass + dimer + penalty
  - Summary:        Per cell-type-config counts (blocks, primers, bowtie pass/fail, dimer)
  - DMR_Blocks:     One row per DMR block with primer count
  - Per cell type:  Individual sheets for each cell-type config

Usage:
    python3 generate_master_excel.py [--results-dir results/] [--output master_primer_assays.xlsx]
"""

import os
import sys
import json
import argparse
from typing import Dict, List
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ── Column definitions ──────────────────────────────────────────────────────

DMR_COLUMNS = [
    "cell_type", "source", "rank", "chrom", "dmr_start", "dmr_end",
    "gene", "annotation", "delta_means", "cleanliness_score",
    "target_mean_meth", "background_mean_meth", "num_cpgs", "block_len",
]

PRIMER_COLUMNS = [
    "assay_id", "seq_id", "template_used", "assay",
    "left_primer", "right_primer",
    "left_primer_cpg_display", "right_primer_cpg_display",
    "left_tm_C", "right_tm_C", "product_size_bp",
    "c_total", "left_c_total", "right_c_total",
    "c_total_tail", "left_c_tail", "right_c_tail",
    "penalty", "left_gc_percent", "right_gc_percent",
    "sense_meth_mismatch_score", "sense_unmeth_mismatch_score",
    "anti_meth_mismatch_score", "anti_unmeth_mismatch_score",
    "bowtie_passes_filter", "bowtie_intended_genome",
    "mismatch_profile", "mapping_error_note",
    "left_structure_mfe_kcal_mol", "right_structure_mfe_kcal_mol",
    "primer_dimer_prediction", "primer_dimer_end_min_dg",
    "common_variant_score",
]

ALL_COLUMNS = DMR_COLUMNS + PRIMER_COLUMNS

COL_WIDTHS = {
    "cell_type": 10, "source": 10, "rank": 6,
    "chrom": 8, "dmr_start": 12, "dmr_end": 12,
    "gene": 12, "annotation": 15, "delta_means": 10,
    "cleanliness_score": 10, "target_mean_meth": 12, "background_mean_meth": 14,
    "num_cpgs": 8, "block_len": 9,
    "assay_id": 18, "seq_id": 14, "template_used": 12, "assay": 6,
    "left_primer": 28, "right_primer": 28,
    "left_primer_cpg_display": 28, "right_primer_cpg_display": 28,
    "left_tm_C": 10, "right_tm_C": 10, "product_size_bp": 12,
    "c_total": 8, "left_c_total": 10, "right_c_total": 10,
    "c_total_tail": 10, "left_c_tail": 8, "right_c_tail": 8,
    "penalty": 8, "left_gc_percent": 10, "right_gc_percent": 10,
    "sense_meth_mismatch_score": 12, "sense_unmeth_mismatch_score": 12,
    "anti_meth_mismatch_score": 12, "anti_unmeth_mismatch_score": 12,
    "bowtie_passes_filter": 10, "bowtie_intended_genome": 18,
    "mismatch_profile": 50, "mapping_error_note": 30,
    "left_structure_mfe_kcal_mol": 12, "right_structure_mfe_kcal_mol": 12,
    "primer_dimer_prediction": 20, "primer_dimer_end_min_dg": 12,
    "common_variant_score": 10,
}

SUMMARY_COLUMNS = [
    "cell_type", "source", "num_blocks", "num_primers",
    "num_pass_bowtie", "num_fail_bowtie", "num_bowtie_none",
    "num_low_dimer", "num_medium_dimer", "num_high_dimer",
]

DMR_BLOCK_COLUMNS = [
    "cell_type", "source", "seq_id", "rank",
    "chrom", "dmr_start", "dmr_end", "gene", "annotation",
    "delta_means", "cleanliness_score", "target_mean_meth",
    "background_mean_meth", "num_cpgs", "block_len",
    "num_primers", "has_bowtie_pass",
]


# ── Data loading ────────────────────────────────────────────────────────────

def _source_label(dir_name: str) -> str:
    if dir_name.endswith("_atlas"):
        return "atlas"
    elif dir_name.endswith("_relaxed"):
        return "relaxed"
    return "v2.2.8"


def _cell_type_from_dir(dir_name: str) -> str:
    for suffix in ("_atlas", "_relaxed"):
        if dir_name.endswith(suffix):
            return dir_name[:-len(suffix)]
    return dir_name


def load_results_dir(results_dir: str) -> List[dict]:
    """Scan all subdirectories, load primers.json + dmr_blocks.json, join on seq_id."""
    all_rows = []

    for entry in sorted(os.listdir(results_dir)):
        dir_path = os.path.join(results_dir, entry)
        if not os.path.isdir(dir_path):
            continue

        primers_path = os.path.join(dir_path, "primers.json")
        blocks_path = os.path.join(dir_path, "dmr_blocks.json")

        if not os.path.exists(primers_path):
            continue

        with open(primers_path) as f:
            primers = json.load(f)
        if not primers:
            continue

        block_lookup: Dict[str, dict] = {}
        if os.path.exists(blocks_path):
            with open(blocks_path) as f:
                blocks = json.load(f)
            for b in blocks:
                block_lookup[b["seq_id"]] = b

        cell_type = _cell_type_from_dir(entry)
        source = _source_label(entry)

        for p in primers:
            row = _build_primer_row(p, block_lookup, cell_type, source)
            all_rows.append(row)

        print(f"  {entry}: {len(primers)} primers, {len(block_lookup)} blocks")

    return all_rows


def _build_primer_row(p: dict, block_lookup: Dict[str, dict],
                      cell_type: str, source: str) -> dict:
    seq_id = p.get("seq_id", "")
    block = block_lookup.get(seq_id, {})

    template = p.get("template_used", "")
    if template in ("SM", "AM"):
        assay = "M"
    elif template in ("SU", "AU"):
        assay = "U"
    else:
        assay = template

    return {
        # DMR-level
        "cell_type": cell_type, "source": source,
        "rank": block.get("rank"), "chrom": block.get("chrom"),
        "dmr_start": block.get("start"), "dmr_end": block.get("end"),
        "gene": block.get("gene"), "annotation": block.get("annotation"),
        "delta_means": block.get("delta_means"),
        "cleanliness_score": block.get("cleanliness_score"),
        "target_mean_meth": block.get("target_mean_meth"),
        "background_mean_meth": block.get("background_mean_meth"),
        "num_cpgs": block.get("num_cpgs"), "block_len": block.get("block_len"),
        # Primer-level
        "assay_id": p.get("assay_id"), "seq_id": seq_id,
        "template_used": template, "assay": assay,
        "left_primer": p.get("left_primer"), "right_primer": p.get("right_primer"),
        "left_primer_cpg_display": p.get("left_primer_display"),
        "right_primer_cpg_display": p.get("right_primer_display"),
        "left_tm_C": round(p["left_tm"], 2) if p.get("left_tm") is not None else None,
        "right_tm_C": round(p["right_tm"], 2) if p.get("right_tm") is not None else None,
        "product_size_bp": p.get("product_size"),
        "c_total": p.get("c_total"), "left_c_total": p.get("left_c_total"),
        "right_c_total": p.get("right_c_total"),
        "c_total_tail": p.get("c_total_tail"),
        "left_c_tail": p.get("left_c_tail"), "right_c_tail": p.get("right_c_tail"),
        "penalty": p.get("penalty"),
        "left_gc_percent": p.get("left_gc_percent"),
        "right_gc_percent": p.get("right_gc_percent"),
        "sense_meth_mismatch_score": p.get("sense_meth_mismatch_score"),
        "sense_unmeth_mismatch_score": p.get("sense_unmeth_mismatch_score"),
        "anti_meth_mismatch_score": p.get("anti_meth_mismatch_score"),
        "anti_unmeth_mismatch_score": p.get("anti_unmeth_mismatch_score"),
        "bowtie_passes_filter": p.get("bowtie_passes_filter"),
        "bowtie_intended_genome": p.get("bowtie_intended_genome"),
        "mismatch_profile": p.get("mismatch_profile"),
        "mapping_error_note": p.get("mapping_error_note"),
        "left_structure_mfe_kcal_mol": p.get("left_structure_mfe"),
        "right_structure_mfe_kcal_mol": p.get("right_structure_mfe"),
        "primer_dimer_prediction": p.get("primer_dimer_prediction"),
        "primer_dimer_end_min_dg": p.get("primer_dimer_end_min_dg"),
        "common_variant_score": p.get("common_variant_score"),
    }


# ── Best-per-DMR selection ──────────────────────────────────────────────────

def select_best_per_dmr(all_rows: List[dict]) -> List[dict]:
    """
    Best primer per DMR block. Ranking:
      1. bowtie_passes_filter (True > False > None)
      2. primer_dimer_prediction (low > medium > high)
      3. lower penalty
    """
    groups = defaultdict(list)
    for row in all_rows:
        key = (row["cell_type"], row["source"], row["seq_id"])
        groups[key].append(row)

    dimer_rank = {"low": 0, "medium": 1, "high": 2}
    bowtie_rank = {True: 0, False: 1, None: 2}

    best_rows = []
    for key, rows in groups.items():
        best = min(rows, key=lambda r: (
            bowtie_rank.get(r.get("bowtie_passes_filter"), 2),
            dimer_rank.get(r.get("primer_dimer_prediction"), 3),
            r.get("penalty") or 999,
        ))
        best_rows.append(best)

    best_rows.sort(key=lambda r: (
        r.get("cell_type", ""), r.get("source", ""), r.get("rank") or 999,
    ))
    return best_rows


# ── Summary + DMR blocks ────────────────────────────────────────────────────

def build_summary(all_rows: List[dict]) -> List[dict]:
    stats = defaultdict(lambda: {
        "num_primers": 0, "num_pass_bowtie": 0, "num_fail_bowtie": 0,
        "num_bowtie_none": 0, "num_low_dimer": 0, "num_medium_dimer": 0,
        "num_high_dimer": 0, "seq_ids": set(),
    })

    for r in all_rows:
        key = (r["cell_type"], r["source"])
        s = stats[key]
        s["num_primers"] += 1
        s["seq_ids"].add(r["seq_id"])

        bt = r.get("bowtie_passes_filter")
        if bt is True:
            s["num_pass_bowtie"] += 1
        elif bt is False:
            s["num_fail_bowtie"] += 1
        else:
            s["num_bowtie_none"] += 1

        dim = r.get("primer_dimer_prediction")
        if dim == "low":
            s["num_low_dimer"] += 1
        elif dim == "medium":
            s["num_medium_dimer"] += 1
        elif dim == "high":
            s["num_high_dimer"] += 1

    rows = []
    for (cell_type, source), s in sorted(stats.items()):
        rows.append({
            "cell_type": cell_type, "source": source,
            "num_blocks": len(s["seq_ids"]), "num_primers": s["num_primers"],
            "num_pass_bowtie": s["num_pass_bowtie"],
            "num_fail_bowtie": s["num_fail_bowtie"],
            "num_bowtie_none": s["num_bowtie_none"],
            "num_low_dimer": s["num_low_dimer"],
            "num_medium_dimer": s["num_medium_dimer"],
            "num_high_dimer": s["num_high_dimer"],
        })
    return rows


def build_dmr_blocks(all_rows: List[dict]) -> List[dict]:
    block_stats = defaultdict(lambda: {"num_primers": 0, "has_bowtie_pass": False})
    block_info = {}

    for r in all_rows:
        key = (r["cell_type"], r["source"], r["seq_id"])
        block_stats[key]["num_primers"] += 1
        if r.get("bowtie_passes_filter") is True:
            block_stats[key]["has_bowtie_pass"] = True
        if key not in block_info:
            block_info[key] = {
                "cell_type": r["cell_type"], "source": r["source"],
                "seq_id": r["seq_id"], "rank": r.get("rank"),
                "chrom": r.get("chrom"),
                "dmr_start": r.get("dmr_start"), "dmr_end": r.get("dmr_end"),
                "gene": r.get("gene"), "annotation": r.get("annotation"),
                "delta_means": r.get("delta_means"),
                "cleanliness_score": r.get("cleanliness_score"),
                "target_mean_meth": r.get("target_mean_meth"),
                "background_mean_meth": r.get("background_mean_meth"),
                "num_cpgs": r.get("num_cpgs"), "block_len": r.get("block_len"),
            }

    rows = []
    for key, info in block_info.items():
        row = dict(info)
        row["num_primers"] = block_stats[key]["num_primers"]
        row["has_bowtie_pass"] = block_stats[key]["has_bowtie_pass"]
        rows.append(row)

    rows.sort(key=lambda r: (
        r.get("cell_type", ""), r.get("source", ""), r.get("rank") or 999,
    ))
    return rows


# ── XLSX writing ────────────────────────────────────────────────────────────

def _style_sheet(ws, columns: List[str], df: pd.DataFrame):
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', text_rotation=45)
        cell.border = thin_border

    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            if col_name == "bowtie_passes_filter" and val is not None:
                cell.fill = pass_fill if val else fail_fill
            if col_name == "primer_dimer_prediction" and val is not None:
                if val == "high":
                    cell.fill = fail_fill
                elif val == "low":
                    cell.fill = pass_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 15)


def write_master_xlsx(all_rows: List[dict], output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: All_Primers
    df_all = pd.DataFrame(all_rows, columns=ALL_COLUMNS)
    ws = wb.create_sheet("All_Primers")
    _style_sheet(ws, ALL_COLUMNS, df_all)
    print(f"  All_Primers: {len(df_all)} rows")

    # Sheet 2: Best_Per_DMR
    best_rows = select_best_per_dmr(all_rows)
    df_best = pd.DataFrame(best_rows, columns=ALL_COLUMNS)
    ws = wb.create_sheet("Best_Per_DMR")
    _style_sheet(ws, ALL_COLUMNS, df_best)
    print(f"  Best_Per_DMR: {len(df_best)} rows")

    # Sheet 3: Summary
    summary_rows = build_summary(all_rows)
    df_summary = pd.DataFrame(summary_rows, columns=SUMMARY_COLUMNS)
    ws = wb.create_sheet("Summary")
    _style_sheet(ws, SUMMARY_COLUMNS, df_summary)
    print(f"  Summary: {len(df_summary)} rows")

    # Sheet 4: DMR_Blocks
    dmr_rows = build_dmr_blocks(all_rows)
    df_dmr = pd.DataFrame(dmr_rows, columns=DMR_BLOCK_COLUMNS)
    ws = wb.create_sheet("DMR_Blocks")
    _style_sheet(ws, DMR_BLOCK_COLUMNS, df_dmr)
    print(f"  DMR_Blocks: {len(df_dmr)} rows")

    # Sheets 5-N: Per cell type
    cell_types = sorted(df_all["cell_type"].unique())
    sources = sorted(df_all["source"].unique())
    for ct in cell_types:
        for src in sources:
            subset = df_all[(df_all["cell_type"] == ct) & (df_all["source"] == src)]
            if len(subset) == 0:
                continue
            sheet_name = f"{ct}_{src}" if src != "v2.2.8" else ct
            sheet_name = sheet_name[:31]
            ws = wb.create_sheet(sheet_name)
            _style_sheet(ws, ALL_COLUMNS, subset)
            print(f"  {sheet_name}: {len(subset)} rows")

    wb.save(output_path)
    print(f"\nMaster Excel saved to {output_path}")


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate master Excel combining all primer assays across cell types"
    )
    parser.add_argument("--results-dir", default="results/",
                        help="Directory containing per-cell-type result subdirectories")
    parser.add_argument("--output", default="master_primer_assays.xlsx",
                        help="Output XLSX path")
    args = parser.parse_args()

    if not os.path.isdir(args.results_dir):
        print(f"Error: results directory '{args.results_dir}' not found")
        sys.exit(1)

    print(f"Scanning {args.results_dir} for primer results...")
    all_rows = load_results_dir(args.results_dir)

    if not all_rows:
        print("Error: no primers found. Run the pipeline first.")
        sys.exit(1)

    print(f"\nTotal: {len(all_rows)} primer pairs across all cell types")
    print(f"\nWriting master Excel to {args.output}...")
    write_master_xlsx(all_rows, args.output)


if __name__ == "__main__":
    main()
