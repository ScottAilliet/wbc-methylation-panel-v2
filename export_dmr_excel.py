#!/usr/bin/env python3
"""
Export DMR regions with per-CpG methylation to Excel.

Reads dmr_blocks.json files from pipeline output directories and produces
a multi-sheet Excel workbook with:
  - One sheet per cell type: top N DMR blocks with per-CpG detail
  - A summary sheet: all blocks across all cell types

Usage:
    # Export from a single cell type:
    python export_dmr_excel.py results/MONO/dmr_blocks.json --top 200

    # Export from all 7 cell types (one sheet each):
    python export_dmr_excel.py results/*/dmr_blocks.json --top 200

    # Custom output path:
    python export_dmr_excel.py results/*/dmr_blocks.json --top 200 -o dmr_regions.xlsx
"""

import sys
import json
import glob
import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Block-level columns (one row per DMR block)
BLOCK_COLUMNS = [
    ("Cell type", "cell_type_id"),
    ("Direction", "direction"),
    ("Rank", "rank"),
    ("Seq ID", "seq_id"),
    ("Chromosome", "chrom"),
    ("Start", "start"),
    ("End", "end"),
    ("Block length (bp)", "block_len"),
    ("# CpGs", "num_cpgs"),
    ("Gene", "gene"),
    ("Annotation", "annotation"),
    ("Target mean methylation", "target_mean_meth"),
    ("Background mean methylation", "background_mean_meth"),
    ("Delta means", "delta_means"),
    ("Cleanliness score", "cleanliness_score"),
]

# Per-CpG columns (one row per CpG within each block)
CPG_COLUMNS = [
    ("Cell type", "cell_type_id"),
    ("Direction", "direction"),
    ("Seq ID", "seq_id"),
    ("Rank", "rank"),
    ("Chromosome", "chrom"),
    ("Block start", "start"),
    ("Block end", "end"),
    ("Gene", "gene"),
    ("CpG label", "label"),
    ("CpG position", "position"),
    ("CpG index", "global_idx"),
    ("Target mean beta", "target_mean_beta"),
    ("Background mean beta", "background_mean_beta"),
    ("Delta beta", "delta_beta"),
]


def load_blocks(json_path: str) -> list:
    """Load DMR blocks from a JSON file."""
    with open(json_path, "r") as f:
        return json.load(f)


def format_block_row(block: dict, subgroup_names: list = None) -> dict:
    """Extract block-level columns from a block dict.

    If subgroup_names is provided, also add per-subgroup methylation columns
    and summary columns (min/max subgroup, worst subgroup).

    For hypomethylated markers (direction='U'): worst = least methylated subgroup.
    For hypermethylated markers (direction='M'): worst = most methylated subgroup.
    """
    row = {}
    for display_name, key in BLOCK_COLUMNS:
        val = block.get(key, "")
        if key == "direction" and not val:
            val = "U"
        row[display_name] = val

    if subgroup_names is not None:
        bg_sg = block.get("bg_subgroup_meth", {})
        is_hyper = block.get("direction", "U") == "M"
        for sg in subgroup_names:
            col_name = f"BG: {sg}"
            row[col_name] = round(bg_sg.get(sg, ""), 4) if sg in bg_sg else ""
        # Summary columns
        if bg_sg:
            if is_hyper:
                # Hyper: worst = most methylated background subgroup
                row["Max subgroup methylation"] = round(max(bg_sg.values()), 4)
                row["Worst background subgroup"] = max(bg_sg, key=bg_sg.get)
            else:
                # Hypo: worst = least methylated background subgroup
                row["Min subgroup methylation"] = round(min(bg_sg.values()), 4)
                row["Worst background subgroup"] = min(bg_sg, key=bg_sg.get)
        else:
            row["Min subgroup methylation"] = ""
            row["Worst background subgroup"] = ""

    return row


def format_cpg_rows(block: dict) -> list:
    """Extract one row per CpG from a block dict."""
    rows = []
    direction = block.get("direction", "U")
    for cpg in block.get("cpg_sites", []):
        row = {}
        for display_name, key in CPG_COLUMNS:
            if key == "direction":
                row[display_name] = direction
            elif key in cpg:
                row[display_name] = cpg[key]
            else:
                row[display_name] = block.get(key, "")
        # Compute delta_beta if not present
        if "Delta beta" not in row or row["Delta beta"] == "":
            tg = cpg.get("target_mean_beta", 0)
            bg = cpg.get("background_mean_beta", 0)
            row["Delta beta"] = round(abs(tg - bg), 4)
        rows.append(row)
    return rows


def write_sheet(ws, headers, rows, freeze="A2"):
    """Write headers and rows to a worksheet with styling."""
    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.border = THIN_BORDER

    # Auto-width (approximate)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, min(len(rows) + 2, 200)):  # sample first 200 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)

    # Freeze header row
    ws.freeze_panes = freeze


def export_excel(json_paths, top_n, output_path):
    """Export DMR blocks from multiple JSON files to a multi-sheet Excel."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # First pass: collect all unique background subgroup names across all files
    all_subgroup_names = set()
    for json_path in sorted(json_paths):
        blocks = load_blocks(json_path)
        for b in blocks:
            bg_sg = b.get("bg_subgroup_meth", {})
            all_subgroup_names.update(bg_sg.keys())
    subgroup_names = sorted(all_subgroup_names)

    all_block_rows = []
    all_cpg_rows = []

    for json_path in sorted(json_paths):
        blocks = load_blocks(json_path)
        if not blocks:
            print(f"  Warning: no blocks in {json_path}, skipping")
            continue

        # Sort by rank (rank 1 = best), take top N
        blocks_sorted = sorted(blocks, key=lambda b: b.get("rank", 999))
        top_blocks = blocks_sorted[:top_n]

        cell_type = top_blocks[0].get("cell_type_id", "Unknown")
        is_hyper = top_blocks[0].get("direction", "U") == "M"
        direction_label = "hyper" if is_hyper else "hypo"
        sheet_name = f"{cell_type}_{direction_label}"[:31]
        print(f"  {json_path}: {len(blocks)} blocks → top {len(top_blocks)} for {cell_type} ({direction_label})")

        # Collect subgroups that actually appear in THIS cell type's data
        ct_subgroups = set()
        for b in top_blocks:
            ct_subgroups.update(b.get("bg_subgroup_meth", {}).keys())
        ct_subgroup_names = sorted(ct_subgroups)

        # Per-cell-type sheet: only show columns for this cell type's background subgroups
        block_rows = [format_block_row(b, ct_subgroup_names) for b in top_blocks]
        block_headers = [name for name, _ in BLOCK_COLUMNS]
        for sg in ct_subgroup_names:
            block_headers.append(f"BG: {sg}")
        # Detect if any blocks are hypermethylated
        has_hyper = any(b.get("direction", "U") == "M" for b in top_blocks)
        has_hypo = any(b.get("direction", "U") != "M" for b in top_blocks)
        if has_hypo:
            block_headers.append("Min subgroup methylation")
        if has_hyper:
            block_headers.append("Max subgroup methylation")
        block_headers.append("Worst background subgroup")
        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, block_headers, block_rows)

        # Collect for summary sheets (use all subgroup names for the summary)
        all_block_rows.extend(
            [format_block_row(b, subgroup_names) for b in top_blocks]
        )
        for b in top_blocks:
            all_cpg_rows.extend(format_cpg_rows(b))

    # Summary sheet: all blocks (with per-subgroup columns)
    if all_block_rows:
        block_headers = [name for name, _ in BLOCK_COLUMNS]
        for sg in subgroup_names:
            block_headers.append(f"BG: {sg}")
        # Include both min and max columns (one will be empty per block)
        block_headers.append("Min subgroup methylation")
        block_headers.append("Max subgroup methylation")
        block_headers.append("Worst background subgroup")
        ws_summary = wb.create_sheet(title="All blocks summary", index=0)
        write_sheet(ws_summary, block_headers, all_block_rows)

    # Per-CpG sheet: all CpGs across all cell types
    if all_cpg_rows:
        cpg_headers = [name for name, _ in CPG_COLUMNS]
        ws_cpg = wb.create_sheet(title="Per-CpG detail")
        write_sheet(ws_cpg, cpg_headers, all_cpg_rows)

    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print(f"  {len(all_block_rows)} blocks across {len(json_paths)} cell type(s)")
    print(f"  {len(all_cpg_rows)} CpG sites total")


def main():
    parser = argparse.ArgumentParser(
        description="Export DMR regions with per-CpG methylation to Excel."
    )
    parser.add_argument(
        "inputs", nargs="+",
        help="Path(s) to dmr_blocks.json file(s). Supports glob patterns."
    )
    parser.add_argument(
        "--top", type=int, default=200,
        help="Number of top DMR regions per cell type (default: 200)"
    )
    parser.add_argument(
        "-o", "--output", default="dmr_regions.xlsx",
        help="Output Excel file path (default: dmr_regions.xlsx)"
    )
    args = parser.parse_args()

    # Expand glob patterns
    json_paths = []
    for pattern in args.inputs:
        matches = sorted(glob.glob(pattern))
        if matches:
            json_paths.extend(matches)
        elif Path(pattern).exists():
            json_paths.append(pattern)
        else:
            print(f"Warning: no files matched '{pattern}'")

    if not json_paths:
        print("Error: no dmr_blocks.json files found.")
        sys.exit(1)

    # Deduplicate
    json_paths = sorted(set(json_paths))

    print(f"Exporting top {args.top} DMR regions from {len(json_paths)} file(s):")
    export_excel(json_paths, args.top, args.output)


if __name__ == "__main__":
    main()
